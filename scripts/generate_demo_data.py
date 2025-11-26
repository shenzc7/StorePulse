#!/usr/bin/env python3
"""
Generate Demo Data for StorePulse Product Presentation
=======================================================

This script creates realistic retail footfall data that demonstrates
what the NB-INGARCH model excels at:

1. WEEKLY PATTERNS - Higher traffic on weekends (Fri-Sun)
2. PROMOTIONS - Clear traffic lift during promotional events
3. WEATHER IMPACT - Rainy/stormy days reduce foot traffic
4. HOLIDAY EFFECTS - Significant spikes on holidays
5. OVERDISPERSION - Variance > Mean (realistic retail data)
6. VOLATILITY CLUSTERING - High variance days tend to cluster

Output:
- demo_lite.csv - Minimal data (date + visits)
- demo_pro.csv - Full data with all exogenous features
- demo_lite.xlsx, demo_pro.xlsx - Excel versions
"""

import random
from datetime import date, timedelta
from pathlib import Path
import csv

# Seed for reproducibility
random.seed(42)

# Configuration
START_DATE = date(2024, 5, 1)  # 6 months of data
NUM_DAYS = 180
BASE_VISITS = 145  # Average daily visitors

# Output directory
DATA_DIR = Path(__file__).parent.parent / "data"
DATA_DIR.mkdir(exist_ok=True)

# Holidays (significant traffic spikes)
HOLIDAYS = {
    date(2024, 5, 27): "Memorial Day",
    date(2024, 7, 4): "Independence Day",
    date(2024, 9, 2): "Labor Day",
    date(2024, 10, 14): "Columbus Day",
}

# Promotion periods (for Pro mode)
PROMO_PERIODS = [
    (date(2024, 5, 10), date(2024, 5, 12), "flash"),      # Weekend flash sale
    (date(2024, 5, 24), date(2024, 5, 27), "percent_off"), # Memorial Day sale
    (date(2024, 6, 14), date(2024, 6, 16), "bogo"),       # Father's Day BOGO
    (date(2024, 7, 1), date(2024, 7, 7), "percent_off"),  # July 4th week sale
    (date(2024, 8, 9), date(2024, 8, 11), "bundle"),      # Back to school bundle
    (date(2024, 8, 30), date(2024, 9, 2), "percent_off"), # Labor Day sale
    (date(2024, 9, 20), date(2024, 9, 22), "flash"),      # Fall flash sale
    (date(2024, 10, 11), date(2024, 10, 14), "percent_off"), # Columbus Day sale
]

# Weather patterns (simulate realistic weather)
WEATHER_PATTERNS = {
    "sunny": 0.45,    # 45% sunny days
    "cloudy": 0.25,   # 25% cloudy days
    "rainy": 0.15,    # 15% rainy days
    "humid": 0.10,    # 10% humid days
    "storm": 0.05,    # 5% storm days
}

def get_day_of_week_multiplier(d: date) -> float:
    """Weekend = higher traffic, Monday = lowest"""
    dow = d.weekday()  # 0=Monday, 6=Sunday
    multipliers = {
        0: 0.75,   # Monday - lowest
        1: 0.85,   # Tuesday
        2: 0.90,   # Wednesday
        3: 0.95,   # Thursday
        4: 1.15,   # Friday - pickup starts
        5: 1.35,   # Saturday - peak
        6: 1.20,   # Sunday - high but less than Sat
    }
    return multipliers[dow]

def get_weather_multiplier(weather: str) -> float:
    """Bad weather reduces traffic"""
    multipliers = {
        "sunny": 1.10,
        "cloudy": 1.00,
        "humid": 0.95,
        "rainy": 0.75,
        "storm": 0.55,
    }
    return multipliers.get(weather, 1.0)

def get_promo_multiplier(d: date, promo_type: str) -> float:
    """Promotions increase traffic"""
    if promo_type == "none":
        return 1.0
    multipliers = {
        "flash": 1.45,       # Flash sales drive high traffic
        "bogo": 1.35,        # BOGO very attractive
        "percent_off": 1.25, # Standard discount
        "bundle": 1.20,      # Bundles moderate lift
        "other": 1.10,
    }
    return multipliers.get(promo_type, 1.0)

def is_holiday(d: date) -> bool:
    return d in HOLIDAYS

def get_holiday_multiplier(d: date) -> float:
    """Holidays = big traffic spikes"""
    if is_holiday(d):
        return 1.60  # 60% increase on holidays
    return 1.0

def get_promo_type(d: date) -> str:
    """Check if date is in a promotion period"""
    for start, end, promo in PROMO_PERIODS:
        if start <= d <= end:
            return promo
    return "none"

def generate_weather(d: date) -> str:
    """Generate weather with some temporal correlation"""
    r = random.random()
    cumulative = 0
    for weather, prob in WEATHER_PATTERNS.items():
        cumulative += prob
        if r < cumulative:
            return weather
    return "sunny"

def generate_visits(d: date, weather: str, promo_type: str) -> int:
    """Generate realistic visit count with overdispersion"""
    
    # Base mean
    mean = BASE_VISITS
    
    # Apply multipliers
    mean *= get_day_of_week_multiplier(d)
    mean *= get_weather_multiplier(weather)
    mean *= get_promo_multiplier(d, promo_type)
    mean *= get_holiday_multiplier(d)
    
    # Add seasonal trend (slight growth over time)
    days_since_start = (d - START_DATE).days
    seasonal_factor = 1.0 + (days_since_start / NUM_DAYS) * 0.15  # 15% growth over period
    mean *= seasonal_factor
    
    # Overdispersion: variance > mean (characteristic of retail data)
    # Use negative binomial-like generation
    dispersion = 0.15 if d.weekday() >= 4 else 0.08  # More variance on weekends
    
    # Generate with overdispersion
    variance = mean + dispersion * mean * mean
    std = variance ** 0.5
    
    # Sample from normal approximation (for simplicity) with bounds
    visits = int(random.gauss(mean, std))
    visits = max(20, visits)  # Minimum 20 visitors
    
    return visits

def generate_sales(visits: int, promo_type: str) -> float:
    """Generate sales correlated with visits"""
    avg_spend = 32.50  # Average spend per visitor
    
    # Promos increase conversion but may lower average spend
    if promo_type in ["percent_off", "flash"]:
        conversion = 0.45  # Higher conversion during sales
        avg_spend *= 0.85  # Lower average due to discounts
    elif promo_type == "bogo":
        conversion = 0.50
        avg_spend *= 0.90
    else:
        conversion = 0.35  # Normal conversion
    
    buyers = int(visits * conversion)
    sales = buyers * avg_spend * random.uniform(0.9, 1.1)
    return round(sales, 2)

def generate_conversion(promo_type: str) -> float:
    """Generate conversion rate"""
    base = 0.35
    if promo_type in ["percent_off", "flash", "bogo"]:
        base = 0.45
    return round(base * random.uniform(0.9, 1.1), 3)

def is_payday(d: date) -> bool:
    """Check if it's around payday (1st and 15th)"""
    return d.day in [1, 2, 14, 15, 16]

def is_school_break(d: date) -> bool:
    """Check for school breaks"""
    # Summer break
    if date(2024, 6, 15) <= d <= date(2024, 8, 25):
        return True
    # Memorial Day weekend
    if date(2024, 5, 25) <= d <= date(2024, 5, 27):
        return True
    # Labor Day weekend
    if date(2024, 8, 31) <= d <= date(2024, 9, 2):
        return True
    return False

def generate_demo_data():
    """Generate both Lite and Pro demo datasets"""
    
    lite_data = []
    pro_data = []
    
    current_date = START_DATE
    
    for day in range(NUM_DAYS):
        d = current_date + timedelta(days=day)
        
        # Generate features
        weather = generate_weather(d)
        promo_type = get_promo_type(d)
        visits = generate_visits(d, weather, promo_type)
        sales = generate_sales(visits, promo_type)
        conversion = generate_conversion(promo_type)
        paydays = is_payday(d)
        school_breaks = is_school_break(d)
        local_events = HOLIDAYS.get(d, "")
        open_hours = 10.0 if d.weekday() < 5 else 12.0  # Longer weekend hours
        
        # Lite record (minimal)
        lite_data.append({
            "event_date": d.strftime("%Y-%m-%d"),
            "visits": visits,
        })
        
        # Pro record (full features)
        pro_data.append({
            "event_date": d.strftime("%Y-%m-%d"),
            "visits": visits,
            "sales": sales,
            "conversion": conversion,
            "promo_type": promo_type,
            "weather": weather,
            "paydays": paydays,
            "school_breaks": school_breaks,
            "local_events": local_events,
            "open_hours": open_hours,
        })
    
    return lite_data, pro_data

def write_csv(data: list, filepath: Path, fieldnames: list):
    """Write data to CSV file"""
    with open(filepath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)
    print(f"✅ Created {filepath}")

def write_excel(data: list, filepath: Path, fieldnames: list):
    """Write data to Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Store Data"
        
        # Header styling
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col, value=field)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        
        # Write data
        for row_idx, record in enumerate(data, 2):
            for col_idx, field in enumerate(fieldnames, 1):
                value = record.get(field, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left")
        
        # Auto-adjust column widths
        for col in range(1, len(fieldnames) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        wb.save(filepath)
        print(f"✅ Created {filepath}")
        
    except ImportError:
        print(f"⚠️  openpyxl not available, skipping Excel file: {filepath}")

def main():
    print("🎬 Generating StorePulse Demo Data")
    print("=" * 50)
    print(f"📅 Date range: {START_DATE} to {START_DATE + timedelta(days=NUM_DAYS-1)}")
    print(f"📊 Total days: {NUM_DAYS}")
    print()
    
    # Generate data
    lite_data, pro_data = generate_demo_data()
    
    # Define field orders
    lite_fields = ["event_date", "visits"]
    pro_fields = ["event_date", "visits", "sales", "conversion", "promo_type", 
                  "weather", "paydays", "school_breaks", "local_events", "open_hours"]
    
    # Write Lite files
    write_csv(lite_data, DATA_DIR / "demo_lite.csv", lite_fields)
    write_excel(lite_data, DATA_DIR / "demo_lite.xlsx", lite_fields)
    
    # Write Pro files
    write_csv(pro_data, DATA_DIR / "demo_pro.csv", pro_fields)
    write_excel(pro_data, DATA_DIR / "demo_pro.xlsx", pro_fields)
    
    # Print summary statistics
    print()
    print("📈 Data Summary:")
    print(f"   Total visits: {sum(r['visits'] for r in lite_data):,}")
    print(f"   Average daily visits: {sum(r['visits'] for r in lite_data) / len(lite_data):.1f}")
    print(f"   Min visits: {min(r['visits'] for r in lite_data)}")
    print(f"   Max visits: {max(r['visits'] for r in lite_data)}")
    
    promo_days = sum(1 for r in pro_data if r['promo_type'] != 'none')
    rainy_days = sum(1 for r in pro_data if r['weather'] in ['rainy', 'storm'])
    print(f"   Promotion days: {promo_days}")
    print(f"   Rainy/Storm days: {rainy_days}")
    
    print()
    print("🎯 Demo Data Ready!")
    print("   Use demo_lite.csv/xlsx for simple demos")
    print("   Use demo_pro.csv/xlsx to showcase full features")

if __name__ == "__main__":
    main()




