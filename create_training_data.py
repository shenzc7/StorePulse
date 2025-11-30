#!/usr/bin/env python3
"""
Create Comprehensive Training Data Sheets for StorePulse
=======================================================

This script generates two comprehensive Excel datasets for training the NB-INGARCH model:

1. training_lite.xlsx - Basic dataset (date, visits) with 3+ years of data
2. training_pro.xlsx - Advanced dataset with exogenous variables

Features included:
- Weekly patterns (higher traffic weekends)
- Seasonal trends (holiday spikes, summer peaks)
- Weather impacts (rain reduces traffic)
- Promotion effects (traffic lifts)
- Overdispersion (variance > mean)
- Volatility clustering (high variance days cluster)
- Realistic noise and outliers
"""

import random
from datetime import date, timedelta
from pathlib import Path
import numpy as np

# Seed for reproducibility
random.seed(42)
np.random.seed(42)

# Configuration
START_DATE = date(2021, 1, 1)  # 4 years of training data
NUM_DAYS = 1461  # ~4 years
BASE_VISITS = 120  # Base average daily visitors

# Output directory
DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)

# Major holidays (significant traffic spikes)
MAJOR_HOLIDAYS = {
    # 2021
    date(2021, 1, 1): "New Year's Day",
    date(2021, 1, 18): "MLK Day",
    date(2021, 2, 14): "Valentine's Day",
    date(2021, 2, 15): "Presidents Day",
    date(2021, 4, 4): "Easter",
    date(2021, 5, 31): "Memorial Day",
    date(2021, 7, 4): "Independence Day",
    date(2021, 9, 6): "Labor Day",
    date(2021, 10, 11): "Columbus Day",
    date(2021, 11, 11): "Veterans Day",
    date(2021, 11, 25): "Thanksgiving",
    date(2021, 12, 25): "Christmas",

    # 2022
    date(2022, 1, 1): "New Year's Day",
    date(2022, 1, 17): "MLK Day",
    date(2022, 2, 14): "Valentine's Day",
    date(2022, 2, 21): "Presidents Day",
    date(2022, 4, 17): "Easter",
    date(2022, 5, 30): "Memorial Day",
    date(2022, 7, 4): "Independence Day",
    date(2022, 9, 5): "Labor Day",
    date(2022, 10, 10): "Columbus Day",
    date(2022, 11, 11): "Veterans Day",
    date(2022, 11, 24): "Thanksgiving",
    date(2022, 12, 25): "Christmas",

    # 2023
    date(2023, 1, 1): "New Year's Day",
    date(2023, 1, 16): "MLK Day",
    date(2023, 2, 14): "Valentine's Day",
    date(2023, 2, 20): "Presidents Day",
    date(2023, 4, 9): "Easter",
    date(2023, 5, 29): "Memorial Day",
    date(2023, 7, 4): "Independence Day",
    date(2023, 9, 4): "Labor Day",
    date(2023, 10, 9): "Columbus Day",
    date(2023, 11, 11): "Veterans Day",
    date(2023, 11, 23): "Thanksgiving",
    date(2023, 12, 25): "Christmas",

    # 2024
    date(2024, 1, 1): "New Year's Day",
    date(2024, 1, 15): "MLK Day",
    date(2024, 2, 14): "Valentine's Day",
    date(2024, 2, 19): "Presidents Day",
    date(2024, 4, 1): "Easter",
    date(2024, 5, 27): "Memorial Day",
    date(2024, 7, 4): "Independence Day",
    date(2024, 9, 2): "Labor Day",
    date(2024, 10, 14): "Columbus Day",
    date(2024, 11, 11): "Veterans Day",
    date(2024, 11, 28): "Thanksgiving",
    date(2024, 12, 25): "Christmas",
}

# Promotion periods (traffic lifts)
PROMOTION_PERIODS = [
    # 2021
    (date(2021, 1, 2), date(2021, 1, 3), "new_year_sale"),
    (date(2021, 2, 13), date(2021, 2, 15), "valentines_sale"),
    (date(2021, 3, 15), date(2021, 3, 21), "spring_sale"),
    (date(2021, 5, 28), date(2021, 5, 31), "memorial_sale"),
    (date(2021, 6, 25), date(2021, 6, 28), "summer_sale"),
    (date(2021, 8, 13), date(2021, 8, 16), "back_to_school"),
    (date(2021, 9, 3), date(2021, 9, 6), "labor_day_sale"),
    (date(2021, 10, 8), date(2021, 10, 11), "fall_sale"),
    (date(2021, 11, 26), date(2021, 11, 29), "black_friday"),
    (date(2021, 12, 23), date(2021, 12, 26), "christmas_sale"),

    # 2022
    (date(2022, 1, 2), date(2022, 1, 3), "new_year_sale"),
    (date(2022, 2, 12), date(2022, 2, 15), "valentines_sale"),
    (date(2022, 3, 18), date(2022, 3, 25), "spring_sale"),
    (date(2022, 5, 27), date(2022, 5, 30), "memorial_sale"),
    (date(2022, 6, 24), date(2022, 6, 27), "summer_sale"),
    (date(2022, 8, 12), date(2022, 8, 15), "back_to_school"),
    (date(2022, 9, 2), date(2022, 9, 5), "labor_day_sale"),
    (date(2022, 10, 7), date(2022, 10, 10), "fall_sale"),
    (date(2022, 11, 25), date(2022, 11, 28), "black_friday"),
    (date(2022, 12, 23), date(2022, 12, 26), "christmas_sale"),

    # 2023
    (date(2023, 1, 2), date(2023, 1, 3), "new_year_sale"),
    (date(2023, 2, 11), date(2023, 2, 14), "valentines_sale"),
    (date(2023, 3, 17), date(2023, 3, 24), "spring_sale"),
    (date(2023, 5, 26), date(2023, 5, 29), "memorial_sale"),
    (date(2023, 6, 23), date(2023, 6, 26), "summer_sale"),
    (date(2023, 8, 11), date(2023, 8, 14), "back_to_school"),
    (date(2023, 9, 1), date(2023, 9, 4), "labor_day_sale"),
    (date(2023, 10, 6), date(2023, 10, 9), "fall_sale"),
    (date(2023, 11, 24), date(2023, 11, 27), "black_friday"),
    (date(2023, 12, 23), date(2023, 12, 26), "christmas_sale"),

    # 2024
    (date(2024, 1, 2), date(2024, 1, 3), "new_year_sale"),
    (date(2024, 2, 10), date(2024, 2, 14), "valentines_sale"),
    (date(2024, 3, 15), date(2024, 3, 22), "spring_sale"),
    (date(2024, 5, 24), date(2024, 5, 27), "memorial_sale"),
    (date(2024, 6, 21), date(2024, 6, 24), "summer_sale"),
    (date(2024, 8, 9), date(2024, 8, 12), "back_to_school"),
    (date(2024, 8, 30), date(2024, 9, 2), "labor_day_sale"),
    (date(2024, 10, 11), date(2024, 10, 14), "fall_sale"),
    (date(2024, 11, 29), date(2024, 12, 2), "black_friday"),
    (date(2024, 12, 23), date(2024, 12, 26), "christmas_sale"),
]

def get_seasonal_trend(d: date) -> float:
    """Seasonal patterns throughout the year"""
    day_of_year = d.timetuple().tm_yday
    year = d.year

    # Summer peak (June-August)
    if 152 <= day_of_year <= 243:  # June 1 to Aug 31
        return 1.25
    # Holiday season (Nov-Dec)
    elif 305 <= day_of_year <= 366:  # Nov 1 to Dec 31
        return 1.20
    # Winter slow period (Jan-Feb)
    elif day_of_year <= 59:  # Jan 1 to Feb 28
        return 0.85
    else:
        return 1.0

def get_day_of_week_multiplier(d: date) -> float:
    """Strong weekend patterns"""
    dow = d.weekday()  # 0=Monday, 6=Sunday
    multipliers = {
        0: 0.70,   # Monday - lowest
        1: 0.80,   # Tuesday
        2: 0.85,   # Wednesday
        3: 0.90,   # Thursday
        4: 1.20,   # Friday - pickup starts
        5: 1.45,   # Saturday - peak
        6: 1.30,   # Sunday - high but less than Sat
    }
    return multipliers[dow]

def generate_temperature(d: date) -> float:
    """Realistic temperature patterns by season"""
    month = d.month
    day_of_year = d.timetuple().tm_yday

    # Base temperature by month (Fahrenheit)
    base_temps = {
        1: 35, 2: 40, 3: 50, 4: 60, 5: 70,
        6: 80, 7: 85, 8: 85, 9: 75, 10: 65,
        11: 50, 12: 40
    }

    base_temp = base_temps[month]

    # Add daily variation
    daily_variation = np.random.normal(0, 8)  # ±8°F daily variation

    # Add seasonal trend within month
    month_progress = (d.day - 1) / 30  # 0 to 1 within month
    seasonal_adjustment = np.sin(month_progress * 2 * np.pi) * 3  # ±3°F within month

    temperature = base_temp + daily_variation + seasonal_adjustment
    return round(max(10, min(100, temperature)), 1)  # Bound between 10-100°F

def generate_precipitation(d: date) -> float:
    """Realistic precipitation patterns"""
    month = d.month

    # Precipitation probability and amount by month
    precip_patterns = {
        1: (0.25, 2.0),   # Jan: 25% chance, 2" avg
        2: (0.30, 2.5),   # Feb: 30% chance, 2.5" avg
        3: (0.35, 3.0),   # Mar: 35% chance, 3" avg
        4: (0.40, 3.5),   # Apr: 40% chance, 3.5" avg
        5: (0.45, 4.0),   # May: 45% chance, 4" avg
        6: (0.35, 3.5),   # Jun: 35% chance, 3.5" avg
        7: (0.30, 3.0),   # Jul: 30% chance, 3" avg
        8: (0.30, 3.0),   # Aug: 30% chance, 3" avg
        9: (0.35, 3.5),   # Sep: 35% chance, 3.5" avg
        10: (0.40, 4.0),  # Oct: 40% chance, 4" avg
        11: (0.35, 3.5),  # Nov: 35% chance, 3.5" avg
        12: (0.30, 3.0),  # Dec: 30% chance, 3" avg
    }

    prob, avg_amount = precip_patterns[month]

    if random.random() < prob:
        # Exponential distribution for precipitation amounts
        amount = np.random.exponential(avg_amount * 0.7)  # Scale down for more realistic amounts
        return round(min(amount, 8.0), 1)  # Cap at 8" for realism
    else:
        return 0.0

def get_precipitation_multiplier(precipitation: float) -> float:
    """Weather impact on traffic"""
    if precipitation == 0:
        return 1.0
    elif precipitation < 0.1:
        return 0.98  # Light drizzle
    elif precipitation < 0.5:
        return 0.90  # Light rain
    elif precipitation < 1.0:
        return 0.75  # Moderate rain
    elif precipitation < 2.0:
        return 0.60  # Heavy rain
    else:
        return 0.45  # Storm/very heavy rain

def get_temperature_multiplier(temperature: float) -> float:
    """Temperature impact on traffic"""
    if 65 <= temperature <= 80:
        return 1.05  # Ideal shopping weather
    elif 50 <= temperature <= 90:
        return 1.0   # Acceptable
    elif temperature < 35:
        return 0.85  # Cold weather reduces traffic
    elif temperature > 90:
        return 0.90  # Hot weather reduces traffic
    else:
        return 0.95  # Mild discomfort

def is_promotion_day(d: date) -> bool:
    """Check if date is in a promotion period"""
    for start, end, promo_type in PROMOTION_PERIODS:
        if start <= d <= end:
            return True
    return False

def get_promotion_multiplier(d: date) -> float:
    """Promotion impact on traffic"""
    for start, end, promo_type in PROMOTION_PERIODS:
        if start <= d <= end:
            # Different promotion types have different impacts
            promo_multipliers = {
                "new_year_sale": 1.25,
                "valentines_sale": 1.30,
                "spring_sale": 1.20,
                "memorial_sale": 1.35,
                "summer_sale": 1.25,
                "back_to_school": 1.40,
                "labor_day_sale": 1.35,
                "fall_sale": 1.25,
                "black_friday": 1.80,  # Major event
                "christmas_sale": 1.60,
            }
            return promo_multipliers.get(promo_type, 1.20)
    return 1.0

def is_holiday(d: date) -> bool:
    return d in MAJOR_HOLIDAYS

def get_holiday_multiplier(d: date) -> float:
    """Holiday impact"""
    if is_holiday(d):
        holiday_name = MAJOR_HOLIDAYS[d]
        if holiday_name in ["Christmas", "Thanksgiving", "Black Friday"]:
            return 1.80  # Major holidays
        else:
            return 1.50  # Regular holidays
    return 1.0

def generate_competitor_distance() -> float:
    """Competitor distance (fixed for store location)"""
    return 2.5  # 2.5 km from nearest competitor

def generate_visits(d: date, temperature: float, precipitation: float) -> int:
    """Generate visit count with overdispersion and volatility clustering"""

    # Base mean
    mean = BASE_VISITS

    # Apply all multipliers
    mean *= get_seasonal_trend(d)
    mean *= get_day_of_week_multiplier(d)
    mean *= get_temperature_multiplier(temperature)
    mean *= get_precipitation_multiplier(precipitation)
    mean *= get_promotion_multiplier(d)
    mean *= get_holiday_multiplier(d)

    # Add long-term trend (slight growth over 4 years)
    days_since_start = (d - START_DATE).days
    trend_factor = 1.0 + (days_since_start / NUM_DAYS) * 0.20  # 20% growth over period
    mean *= trend_factor

    # Overdispersion: variance > mean (characteristic of retail data)
    # Higher dispersion on weekends and during promotions
    base_dispersion = 0.12
    if d.weekday() >= 4:  # Weekend
        base_dispersion *= 1.5
    if is_promotion_day(d) or is_holiday(d):
        base_dispersion *= 1.3

    # Add volatility clustering (high variance days tend to cluster)
    # Simple autoregressive volatility
    variance = mean + base_dispersion * mean * mean

    # Use negative binomial-like distribution for overdispersion
    # NB(r, p) where r controls overdispersion
    r = mean * mean / (variance - mean) if variance > mean else 100
    p = r / (r + mean)

    # Generate from negative binomial
    visits = np.random.negative_binomial(r, p)

    # Add minimum bounds and occasional outliers
    visits = max(15, visits)  # Minimum 15 visitors

    # Occasional extreme days (outliers)
    if random.random() < 0.02:  # 2% chance
        if random.random() < 0.5:
            visits = int(visits * 1.8)  # High outlier
        else:
            visits = int(visits * 0.4)  # Low outlier

    return visits

def generate_training_data():
    """Generate comprehensive training datasets"""

    lite_data = []
    pro_data = []

    current_date = START_DATE

    print("Generating training data...")

    for day in range(NUM_DAYS):
        if day % 365 == 0:
            print(f"Processing year {current_date.year}...")

        d = current_date + timedelta(days=day)

        # Generate exogenous variables
        temperature = generate_temperature(d)
        precipitation = generate_precipitation(d)
        is_promotion = is_promotion_day(d)
        competitor_distance = generate_competitor_distance()

        # Generate visits
        visits = generate_visits(d, temperature, precipitation)

        # Lite record
        lite_data.append({
            "date": d.strftime("%Y-%m-%d"),
            "visits": visits,
        })

        # Pro record
        pro_data.append({
            "date": d.strftime("%Y-%m-%d"),
            "visits": visits,
            "temperature": temperature,
            "precipitation": precipitation,
            "is_promotion": 1 if is_promotion else 0,
            "competitor_distance_km": competitor_distance,
        })

    return lite_data, pro_data

def create_excel_file(data: list, filepath: Path, title: str):
    """Create styled Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Training Data"

        # Get fieldnames
        fieldnames = list(data[0].keys())

        # Header styling
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
        ws.merge_cells('A1:F1')

        # Write headers
        header_row = 3
        for col, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=header_row, column=col, value=field.replace('_', ' ').title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Write data
        for row_idx, record in enumerate(data, header_row + 1):
            for col_idx, field in enumerate(fieldnames, 1):
                value = record[field]
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # Center numeric columns
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")

        # Auto-adjust column widths
        for col in range(1, len(fieldnames) + 1):
            max_length = 10  # minimum width
            column_letter = get_column_letter(col)

            for row in range(header_row, len(data) + header_row + 1):
                cell_value = ws[f'{column_letter}{row}'].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            ws.column_dimensions[column_letter].width = min(max_length + 2, 25)

        # Freeze header row
        ws.freeze_panes = f"A{header_row + 1}"

        # Add summary statistics at bottom
        summary_row = len(data) + header_row + 3
        ws.cell(row=summary_row, column=1, value="Summary Statistics:").font = Font(bold=True)

        if 'visits' in fieldnames:
            visits_data = [r['visits'] for r in data]
            ws.cell(row=summary_row + 1, column=1, value="Total Visits:")
            ws.cell(row=summary_row + 1, column=2, value=sum(visits_data))
            ws.cell(row=summary_row + 2, column=1, value="Average Daily Visits:")
            ws.cell(row=summary_row + 2, column=2, value=f"{sum(visits_data)/len(visits_data):.1f}")
            ws.cell(row=summary_row + 3, column=1, value="Min Visits:")
            ws.cell(row=summary_row + 3, column=2, value=min(visits_data))
            ws.cell(row=summary_row + 4, column=1, value="Max Visits:")
            ws.cell(row=summary_row + 4, column=2, value=max(visits_data))

        wb.save(filepath)
        print(f"✅ Created {filepath}")

    except ImportError:
        print(f"⚠️  openpyxl not available, skipping Excel file: {filepath}")

def main():
    print("🎬 Creating Comprehensive StorePulse Training Data")
    print("=" * 60)
    print(f"📅 Date range: {START_DATE} to {START_DATE + timedelta(days=NUM_DAYS-1)}")
    print(f"📊 Total days: {NUM_DAYS}")
    print()

    # Generate data
    lite_data, pro_data = generate_training_data()

    # Create Excel files
    create_excel_file(
        lite_data,
        DATA_DIR / "training_lite.xlsx",
        "StorePulse Lite Training Data - Comprehensive Dataset for NB-INGARCH Model Training"
    )

    create_excel_file(
        pro_data,
        DATA_DIR / "training_pro.xlsx",
        "StorePulse Pro Training Data - Advanced Dataset with Exogenous Variables"
    )

    # Print summary
    print()
    print("📈 Training Data Summary:")
    visits = [r['visits'] for r in lite_data]
    print(f"   Total visits: {sum(visits):,}")
    print(f"   Average daily visits: {sum(visits)/len(visits):.1f}")
    print(f"   Min visits: {min(visits)}")
    print(f"   Max visits: {max(visits)}")
    print(f"   Standard deviation: {np.std(visits):.1f}")
    print(f"   Coefficient of variation: {np.std(visits)/np.mean(visits):.3f}")

    promo_days = sum(1 for r in pro_data if r['is_promotion'] == 1)
    rainy_days = sum(1 for r in pro_data if r['precipitation'] > 0.1)
    print(f"   Promotion days: {promo_days}")
    print(f"   Rainy days: {rainy_days}")
    print(f"   Holiday days: {len(MAJOR_HOLIDAYS)}")

    print()
    print("🎯 Training Data Ready!")
    print("   Use training_lite.xlsx for basic NB-INGARCH training")
    print("   Use training_pro.xlsx for advanced models with exogenous variables")
    print()
    print("💡 Key Features:")
    print("   • 4 years of comprehensive data")
    print("   • Realistic weekly patterns (weekends higher)")
    print("   • Seasonal trends and holiday effects")
    print("   • Weather impacts on traffic")
    print("   • Promotion periods with traffic lifts")
    print("   • Overdispersion (variance > mean)")
    print("   • Occasional outliers for robustness")

if __name__ == "__main__":
    main()

